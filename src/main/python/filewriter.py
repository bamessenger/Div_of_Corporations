from openpyxl import Workbook, load_workbook


class WriteFile:
    def __init__(self, fPath):
        self.currMaster = fPath
        self.wb = Workbook()
        self.existingDict = {}

    def currExcel2Dict(self):
        # load workbook
        self.wb = load_workbook(self.currMaster)
        # select workbook
        sheet = self.wb.active
        # get max row count
        max_row = sheet.max_row
        # iterate through all rows in first column, skipping first row
        for r in range(2, max_row):
            key = sheet.cell(row=r, column=1)
            entityName = sheet.cell(row=r, column=2)
            incDate = sheet.cell(row=r, column=3)
            self.existingDict[key.value] = entityName, incDate

    def newMast2Excel(self):
        pass

    def dictMerge(self, dict2):
        self.newMast = self.existingDict.update(dict2)

    def setNewMastDict(self, newDict):
        # grab current excel data and create a temporary dictionary
        self.currExcel2Dict()
        # receive scraped data dictionary and pass it to merging method
        self.dictMerge(newDict)
        # once both dictionaries are merged, re-write the updated dictionary
        # back to Excel file
        self.newMast2Excel()
